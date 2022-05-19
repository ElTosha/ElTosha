import re
import time
import openpyxl

f = open('sen.txt')
f.readline()

FILENAME = "temp.xlsx"
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "ONT Temperature"
workbook.save(FILENAME)

sheet.cell(row=1, column=1).value = "Count"
sheet.cell(row=1, column=2).value = "Air temp"
sheet.cell(row=1, column=3).value = "Up Temp"
sheet.cell(row=1, column=4).value = "Down Temp"

chart = openpyxl.chart.LineChart()
chart.title = "ONT Temperature"
chart.style = 12
chart.x_axis.title = 'Count'
chart.y_axis.title = 'Temp'
#chart.x_axis.number_format = '0'
chart.height = 20
chart.width = 40

workbook.save(FILENAME)

workbook = openpyxl.load_workbook(FILENAME)
sheet = workbook.active
sheet.title = "ONT Temperature"

result = {}
sens1 = []
sens2 = []
sens3 = []

for i in f:
	sensor_tem = re.search('\d+\s+(28\S+)\s+(\S+)', i)
	if sensor_tem: 	
		sens = sensor_tem.group(1)
		temp = sensor_tem.group(2)
		if sens == '28:3D:A9:CE:4E:20:01:0F':
			sens1.append(temp)
		if sens == '28:EF:51:CC:4E:20:01:3A':
			sens2.append(temp)
		if sens == '28:3F:AC:1E:4F:20:01:19':
			sens3.append(temp)


for i in range(1,200):
	sheet.cell(row=i+1, column=1).value = i

result = {'Up':sens1, 'Down':sens2, 'Air':sens3}
print(result)
for key, value in result.items():
	print("{0}: {1}".format(key,value))

	
for key, znach in result.items():
	if key == 'Up':
		stroka = 2
		for i in znach:
			sheet.cell(row=stroka, column=2).value = float(i)
			stroka += 1
	if key == 'Down':
		stroka = 2
		for i in znach:
			sheet.cell(row=stroka, column=3).value = float(i)
			stroka += 1
	elif key == 'Air':
		stroka = 2
		for i in znach:
			sheet.cell(row=stroka, column=4).value = float(i)
			stroka += 1


workbook.save(FILENAME)


for j in range (2, 5):
	values = openpyxl.chart.Reference(sheet, min_col=j, min_row=1, max_row=50)
	chart.add_data(values, titles_from_data=True)
xvalues = openpyxl.chart.Reference(sheet, min_col=1, min_row=2, max_row=50)
chart.set_categories(xvalues)
sheet.add_chart(chart, "F2")
workbook.save(FILENAME)